from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_

from app.models.schedule import Schedule
from app.repositories.schedule import ScheduleRepository
from app.schemas.schedule import ScheduleIn, ScheduleExportParams

from app.services.base import BaseService

import io
from datetime import datetime, timedelta
from typing import List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# PDF imports
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm

from ..models import Lesson, Group, StudyForm


class ScheduleService(BaseService[Schedule, ScheduleIn]):
    """
    Service layer for schedule management.

    Responsibilities:
    - Create schedules with uniqueness validation per (semester, direction).
    - Export schedules in Excel or PDF formats.
    - Provide helper routines for building tabular representations (time slots, table data, styling).
    """

    def __init__(self, db: Session):
        """
        Initialize the Schedule service.

        Args:
            db (Session): Active SQLAlchemy session.
        """
        super().__init__(db, Schedule, ScheduleRepository(db))

    def create(self, schedule: ScheduleIn) -> Schedule:
        """
        Create a new schedule ensuring uniqueness within the semester and direction.

        Args:
            schedule (ScheduleIn): Payload for the schedule to create.

        Returns:
            Schedule: Newly created schedule.

        Raises:
            HTTPException: 400 if a schedule for the same semester and direction already exists.
        """
        found_schedule = (
            self.db.query(Schedule)
            .filter(
                Schedule.semester_id == schedule.semester_id,
                Schedule.direction_id == schedule.direction_id,
            )
            .first()
        )

        if found_schedule:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Schedule for this semester already exists.",
            )

        return super().create(schedule)

    def export(
        self,
        schedule_id: int,
        export_params: ScheduleExportParams,
    ) -> StreamingResponse:
        """
        Export a schedule to the requested format (Excel or PDF) and return a streaming response.

        The method delegates to the specific exporter based on export_params.format,
        builds an appropriate content type, and returns a StreamingResponse with
        download headers.

        Args:
            schedule_id (int): ID of the schedule to export.
            export_params (ScheduleExportParams): Export options (format, group_ids, filename).

        Returns:
            StreamingResponse: Streamed file content with proper headers for download.
        """
        """
        Universal schedule export method with error handling
        Returns a ready StreamingResponse for return from the route
        """
        # Вызываем соответствующий метод экспорта
        if export_params.format == "excel":
            file_buffer, generated_filename = self.export_schedule_excel(
                schedule_id, export_params.group_ids, export_params.filename
            )
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:  # pdf
            file_buffer, generated_filename = self.export_schedule_pdf(
                schedule_id, export_params.group_ids, export_params.filename
            )
            media_type = "application/pdf"

        # Формируем заголовки
        headers = {
            "Content-Disposition": f'attachment; filename="{generated_filename}"',
            "Content-Type": media_type,
        }

        # Возвращаем StreamingResponse
        return StreamingResponse(
            file_buffer,
            media_type=media_type,
            headers=headers,
        )

    def export_schedule_excel(
        self,
        schedule_id: int,
        group_ids: Optional[List[int]] = None,
        custom_filename: Optional[str] = None,
    ) -> Tuple[io.BytesIO, str]:
        """
        Export a schedule to Excel in the tabular layout used by the UI.

        Steps:
        - Load schedule and lessons (optionally filtered by groups).
        - Generate time slots (08:00–18:00, 45-minute slots with 10-minute breaks).
        - Prepare styles and fill the sheet.
        - Return an in-memory buffer and a safe filename.

        Args:
            schedule_id (int): ID of the schedule to export.
            group_ids (Optional[List[int]]): Subset of group IDs to include; None for all.
            custom_filename (Optional[str]): Base filename without extension; defaults to a generated name.

        Returns:
            Tuple[io.BytesIO, str]: (buffer, filename) where buffer is positioned at start.
        """
        """Экспорт расписания в Excel в формате как на картинке"""

        # Получаем расписание
        schedule = self.get_by_id(schedule_id)

        # Получаем уроки для расписания (с фильтрацией по группам если указаны)
        lessons = self._get_lessons_for_schedule(schedule_id, group_ids)

        # Создаем временные слоты (с 8:00 до 18:00 с интервалом 45 минут)
        time_slots = self._generate_time_slots()

        # Создаем Excel файл
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Schedule"

        # Определяем период дат из уроков
        if lessons:
            start = min(lesson.date for lesson in lessons)
            end = max(lesson.date for lesson in lessons)
        else:
            # Если уроков нет, используем текущую дату
            from datetime import date

            start = end = date.today()

        # Настройка стилей
        self._setup_styles(worksheet, schedule, start, end, time_slots, [], lessons)

        # Генерируем имя файла
        if custom_filename:
            filename = f"{custom_filename}.xlsx"
        else:
            filename = f"schedule_{schedule.direction.name}_{schedule.semester.number}_{start.strftime('%Y-%m-%d')}_to_{end.strftime('%Y-%m-%d')}.xlsx"
        filename = filename.replace(" ", "_").replace("/", "_")

        # Сохранение в BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output, filename

    def export_schedule_pdf(
        self,
        schedule_id: int,
        group_ids: Optional[List[int]] = None,
        custom_filename: Optional[str] = None,
    ) -> Tuple[io.BytesIO, str]:
        """
        Export a schedule to PDF using the same data layout as Excel (landscape A4).

        Steps:
        - Load schedule and lessons (optionally filtered by groups).
        - Generate time slots and determine date range.
        - Build a ReportLab table with compact styling and colored cells.

        Args:
            schedule_id (int): ID of the schedule to export.
            group_ids (Optional[List[int]]): Subset of group IDs to include; None for all.
            custom_filename (Optional[str]): Base filename without extension; defaults to a generated name.

        Returns:
            Tuple[io.BytesIO, str]: (buffer, filename) where buffer is positioned at start.
        """
        """Экспорт расписания в PDF в том же формате что и Excel"""

        # Получаем расписание

        schedule = self.get_by_id(schedule_id)

        # Получаем уроки для расписания (с фильтрацией по группам если указаны)
        lessons = self._get_lessons_for_schedule(schedule_id, group_ids)

        # Создаем временные слоты (с 8:00 до 18:00 с интервалом 45 минут)
        time_slots = self._generate_time_slots()

        # Определяем период дат из уроков
        if lessons:
            start = min(lesson.date for lesson in lessons)
            end = max(lesson.date for lesson in lessons)
        else:
            # Если уроков нет, используем текущую дату
            from datetime import date

            start = end = date.today()

        # Создаем PDF
        output = io.BytesIO()

        # Используем landscape ориентацию для лучшего размещения таблицы
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        # Создаем содержимое PDF
        story = []

        # Заголовок (оптимизирован для landscape)
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=getSampleStyleSheet()["Heading1"],
            fontSize=14,  # Уменьшенный размер
            alignment=1,  # CENTER
            spaceAfter=8,  # Уменьшенный отступ
            fontName="Helvetica-Bold",
        )

        title_text = f"SCHEDULE PLAN - ACADEMIC YEAR {schedule.academic_year.name}"
        story.append(Paragraph(title_text, title_style))

        # Подзаголовок (оптимизирован для landscape)
        subtitle_style = ParagraphStyle(
            "CustomSubtitle",
            parent=getSampleStyleSheet()["Heading2"],
            fontSize=10,  # Уменьшенный размер
            alignment=1,  # CENTER
            spaceAfter=12,  # Уменьшенный отступ
            fontName="Helvetica-Bold",
        )

        subtitle_text = f"{schedule.direction.name} - {schedule.direction.faculty.name} - ACADEMIC YEAR {schedule.semester.number} SEMESTER"
        story.append(Paragraph(subtitle_text, subtitle_style))

        # Создаем таблицу данных
        table_data, cell_colors = self._create_pdf_table_data(
            start, end, time_slots, lessons
        )

        # Создаем таблицу
        table = Table(table_data)

        # Стили таблицы (оптимизировано для landscape)
        table_style = TableStyle(
            [
                # Границы
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                # Заголовки временных слотов
                ("BACKGROUND", (1, 0), (-2, 0), colors.lightgrey),
                ("ALIGN", (1, 0), (-2, 0), "CENTER"),
                ("VALIGN", (1, 0), (-2, 0), "MIDDLE"),
                ("FONTNAME", (1, 0), (-2, 0), "Helvetica-Bold"),
                ("FONTSIZE", (1, 0), (-2, 0), 6),  # Уменьшенный размер шрифта
                # Колонка GRUPA
                ("BACKGROUND", (-1, 0), (-1, 0), colors.lightgrey),
                ("ALIGN", (-1, 0), (-1, 0), "CENTER"),
                ("VALIGN", (-1, 0), (-1, 0), "MIDDLE"),
                ("FONTNAME", (-1, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (-1, 0), (-1, 0), 7),
                # Колонка дат
                ("BACKGROUND", (0, 1), (0, -1), colors.whitesmoke),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("VALIGN", (0, 1), (0, -1), "MIDDLE"),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 1), (0, -1), 6),  # Уменьшенный размер шрифта
                # Ячейки уроков
                ("ALIGN", (1, 1), (-2, -1), "CENTER"),
                ("VALIGN", (1, 1), (-2, -1), "MIDDLE"),
                ("FONTNAME", (1, 1), (-2, -1), "Helvetica"),
                ("FONTSIZE", (1, 1), (-2, -1), 5),  # Уменьшенный размер шрифта
                # Колонка групп
                ("ALIGN", (-1, 1), (-1, -1), "CENTER"),
                ("VALIGN", (-1, 1), (-1, -1), "MIDDLE"),
                ("FONTNAME", (-1, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (-1, 1), (-1, -1), 6),  # Уменьшенный размер шрифта
            ]
        )

        table.setStyle(table_style)

        # Применяем цвета к ячейкам с уроками
        for (row, col), hex_color in cell_colors.items():
            rgb_color = self._hex_to_rgb(hex_color)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (col, row), (col, row), rgb_color),
                        ("TEXTCOLOR", (col, row), (col, row), colors.white),
                    ]
                )
            )

        # Устанавливаем ширину колонок (оптимизировано для landscape A4)
        available_width = landscape(A4)[0] - 20 * mm  # Общая ширина минус отступы
        date_col_width = 25 * mm  # Колонка дат
        group_col_width = 20 * mm  # Колонка групп

        # Распределяем оставшуюся ширину между временными слотами
        remaining_width = available_width - date_col_width - group_col_width
        time_slot_width = remaining_width / len(time_slots)

        # Если слишком много временных слотов, уменьшаем ширину колонок
        if len(time_slots) > 10:
            # Уменьшаем колонки дат и групп для экономии места
            date_col_width = 20 * mm
            group_col_width = 15 * mm
            remaining_width = available_width - date_col_width - group_col_width
            time_slot_width = remaining_width / len(time_slots)

        col_widths = [date_col_width]  # Колонка дат
        col_widths.extend([time_slot_width] * len(time_slots))  # Временные слоты
        col_widths.append(group_col_width)  # Колонка групп
        table._argW = col_widths

        story.append(table)

        # Генерируем PDF
        doc.build(story)

        # Генерируем имя файла
        if custom_filename:
            filename = f"{custom_filename}.pdf"
        else:
            filename = f"schedule_{schedule.direction.name}_{schedule.semester.number}_{start.strftime('%Y-%m-%d')}_to_{end.strftime('%Y-%m-%d')}.pdf"
        filename = filename.replace(" ", "_").replace("/", "_")

        output.seek(0)
        return output, filename

    def _create_pdf_table_data(self, start_date, end_date, time_slots, lessons):
        """
        Build PDF table data and a map of cell colors.

        The returned table_data is a list of rows for ReportLab Table:
        - Row 0 is the header: ["Date/Time", ...time_slots..., "GRUPA"]
        - Subsequent rows represent days (one or multiple rows depending on groups)

        Args:
            start_date (date): Start date of the range.
            end_date (date): End date of the range.
            time_slots (list[str]): Time slot labels (e.g., "08:00-08:45").
            lessons (list[Lesson]): Lessons to render.

        Returns:
            tuple[list[list[str]], dict[tuple[int, int], str]]:
                (table_data, cell_colors) where cell_colors keys are (row_index, col_index)
                and values are hex colors for background.
        """
        """Создание данных для PDF таблицы"""
        table_data = []
        cell_colors = {}  # Словарь для хранения цветов ячеек: {(row, col): color}

        # Заголовочная строка
        header_row = ["Date/Time"] + time_slots + ["GRUPA"]
        table_data.append(header_row)

        current_date = start_date
        table_row = 1  # Начинаем с 1, так как 0-я строка - заголовок

        while current_date <= end_date:
            # Получаем уроки на этот день
            day_lessons = [lesson for lesson in lessons if lesson.date == current_date]

            if not day_lessons:
                # Пустая строка для дня без уроков
                date_text = f"{self._get_weekday_english(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                row = [date_text] + [""] * len(time_slots) + [""]
                table_data.append(row)
                table_row += 1
            else:
                # Определяем уникальные группы для этого дня
                groups_for_day = {}
                for lesson in day_lessons:
                    if lesson.group:
                        group_name = lesson.group.name
                        if group_name not in groups_for_day:
                            groups_for_day[group_name] = []
                        groups_for_day[group_name].append(lesson)

                if len(groups_for_day) <= 1:
                    # Одна группа или уроки без группы
                    date_text = f"{self._get_weekday_english(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                    row = [date_text]

                    # Заполняем временные слоты
                    for col_idx, time_slot in enumerate(time_slots):
                        lesson_in_slot = self._find_lesson_in_slot(
                            day_lessons, time_slot
                        )
                        if lesson_in_slot:
                            cell_text = self._format_lesson_for_pdf(lesson_in_slot)
                            # Сохраняем цвет для этой ячейки
                            if lesson_in_slot.subject.color:
                                cell_colors[(table_row, col_idx + 1)] = (
                                    lesson_in_slot.subject.color
                                )
                        else:
                            cell_text = ""
                        row.append(cell_text)

                    # Колонка групп
                    groups_text = (
                        "\n".join(sorted(groups_for_day.keys()))
                        if groups_for_day
                        else ""
                    )
                    row.append(groups_text)
                    table_data.append(row)
                    table_row += 1
                else:
                    # Несколько групп - отдельные строки для каждой группы
                    first_group = True
                    for group_name in sorted(groups_for_day.keys()):
                        group_lessons = groups_for_day[group_name]

                        if first_group:
                            date_text = f"{self._get_weekday_english(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                            first_group = False
                        else:
                            date_text = ""

                        row = [date_text]

                        # Заполняем временные слоты для этой группы
                        for col_idx, time_slot in enumerate(time_slots):
                            lesson_in_slot = self._find_lesson_in_slot(
                                group_lessons, time_slot
                            )
                            if lesson_in_slot:
                                cell_text = self._format_lesson_for_pdf(lesson_in_slot)
                                # Сохраняем цвет для этой ячейки
                                if lesson_in_slot.subject.color:
                                    cell_colors[(table_row, col_idx + 1)] = (
                                        lesson_in_slot.subject.color
                                    )
                            else:
                                cell_text = ""
                            row.append(cell_text)

                        # Колонка групп
                        row.append(group_name)
                        table_data.append(row)
                        table_row += 1

            current_date += timedelta(days=1)

        return table_data, cell_colors

    def _hex_to_rgb(self, hex_color: str):
        """
        Convert a hex color string into a ReportLab Color (0..1 per channel).

        Args:
            hex_color (str): Hex color in the form "#RRGGBB" or "RRGGBB".

        Returns:
            reportlab.lib.colors.Color: RGB color for ReportLab.
        """
        """Конвертация hex цвета в RGB для ReportLab"""
        # Убираем # если есть
        hex_color = hex_color.lstrip("#")

        # Конвертируем в RGB (значения от 0 до 1)
        r = int(hex_color[0:2], 16) / 255.0
        g = int(hex_color[2:4], 16) / 255.0
        b = int(hex_color[4:6], 16) / 255.0

        return colors.Color(r, g, b)

    def _format_lesson_for_pdf(self, lesson) -> str:
        """
        Produce a compact, multi-line text representation of a lesson for PDF cells.

        Format:
        - First line: subject name (truncated)
        - Second line: professor (e.g., "A. Surname", truncated)
        - Third line: delivery ("ONLINE" or "R:<room>")

        Args:
            lesson (Lesson): Lesson to render.

        Returns:
            str: Compact string for the PDF cell.
        """
        """Форматирование урока для PDF (компактная версия)"""
        # Сокращаем имена для экономии места
        subject_name = lesson.subject.name
        if len(subject_name) > 15:
            subject_name = subject_name[:12] + "..."

        professor_name = f"{lesson.professor.name[:1]}. {lesson.professor.surname}"
        if len(professor_name) > 15:
            professor_name = professor_name[:12] + "..."

        lesson_text = f"{subject_name}\n{professor_name}\n"

        if lesson.is_online:
            lesson_text += "ONLINE"
        elif lesson.room:
            lesson_text += f"R:{lesson.room.number}"

        return lesson_text

    def _get_lessons_for_schedule(
        self, schedule_id: int, group_ids: Optional[List[int]] = None
    ) -> List[Lesson]:
        """
        Fetch lessons for a schedule, optionally restricted to specific groups.

        Args:
            schedule_id (int): Schedule ID.
            group_ids (Optional[List[int]]): List of group IDs to include; None for all.

        Returns:
            List[Lesson]: Lessons ordered by date and start time.
        """
        """Получение уроков для расписания с возможной фильтрацией по группам"""
        query = self.db.query(Lesson).filter(Lesson.schedule_id == schedule_id)

        # Если указаны конкретные группы, фильтруем по ним
        if group_ids:
            query = query.filter(Lesson.group_id.in_(group_ids))

        return query.order_by(Lesson.date, Lesson.start_time).all()

    def _get_groups_for_schedule(
        self, schedule_id: int, group_ids: Optional[List[int]] = None
    ) -> List[Group]:
        """
        Fetch groups for a schedule, optionally restricted to specific groups.

        Uses the schedule's direction and semester to scope groups.

        Args:
            schedule_id (int): Schedule ID.
            group_ids (Optional[List[int]]): List of group IDs to include; None for all.

        Returns:
            List[Group]: Groups ordered by name.
        """
        """Получение групп для расписания с возможной фильтрацией"""
        schedule = self.db.query(Schedule).filter(Schedule.id == schedule_id).first()

        query = (
            self.db.query(Group)
            .join(StudyForm)
            .filter(
                and_(
                    StudyForm.direction_id == schedule.direction_id,
                    Group.semester_id == schedule.semester_id,
                )
            )
        )

        # Если указаны конкретные группы, фильтруем по ним
        if group_ids:
            query = query.filter(Group.id.in_(group_ids))

        return query.order_by(Group.name).all()

    def _generate_time_slots(self) -> List[str]:
        """
        Generate time slot labels from 08:00 to 18:00.

        Each slot is 45 minutes followed by a 10-minute break, e.g., "08:00-08:45".

        Returns:
            List[str]: Sequence of time slot labels.
        """
        """Генерация временных слотов"""
        slots = []
        start_time = datetime.strptime("08:00", "%H:%M")
        end_time = datetime.strptime("18:00", "%H:%M")

        current = start_time
        while current < end_time:
            next_time = current + timedelta(minutes=45)
            slot = f"{current.strftime('%H:%M')}-{next_time.strftime('%H:%M')}"
            slots.append(slot)
            current = next_time + timedelta(minutes=10)  # 10 минут перерыв

        return slots

    def _setup_styles(
        self, worksheet, schedule, start_date, end_date, time_slots, groups, lessons
    ):
        """
        Apply styles and fill the Excel worksheet with schedule content.

        Writes:
        - Title and subtitle (merged cells)
        - Time slot headers
        - Rows for each date (and per-group rows if multiple groups on the same day)
        - Lesson cells with subject color backgrounds and wrapped text

        Args:
            worksheet (openpyxl.worksheet.worksheet.Worksheet): Active sheet.
            schedule (Schedule): Schedule metadata (title/subtitle context).
            start_date (date): Start date for rendering.
            end_date (date): End date for rendering.
            time_slots (list[str]): Column headers for time intervals.
            groups (list[Group]): Not used in current implementation (reserved).
            lessons (list[Lesson]): Lessons to render in the grid.

        Returns:
            None
        """
        """Настройка стилей и заполнение данных"""

        # Стили
        header_font = Font(bold=True, size=12, color="000000")
        title_font = Font(bold=True, size=14, color="000000")
        cell_font = Font(size=10, color="000000")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Заголовок (строка 1)
        title_text = f"SCHEDULE PLAN - ACADEMIC YEAR {schedule.academic_year.name}"
        worksheet.merge_cells(f"A1:{get_column_letter(len(time_slots) + 2)}1")
        title_cell = worksheet.cell(row=1, column=1, value=title_text)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Подзаголовок (строка 2)
        subtitle_text = f"{schedule.direction.name} - {schedule.direction.faculty.name} - ACADEMIC YEAR {schedule.semester.number} SEMESTER"
        worksheet.merge_cells(f"A2:{get_column_letter(len(time_slots) + 2)}2")
        subtitle_cell = worksheet.cell(row=2, column=1, value=subtitle_text)
        subtitle_cell.font = header_font
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Временные слоты (строка 3)
        for col, time_slot in enumerate(time_slots, 2):
            cell = worksheet.cell(row=3, column=col, value=time_slot)
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Колонка "GRUPA" (строка 3, последняя колонка)
        group_col = len(time_slots) + 2
        group_cell = worksheet.cell(row=3, column=group_col, value="GRUPA")
        group_cell.font = header_font
        group_cell.border = border
        group_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Заполнение дней и уроков
        current_date = start_date
        row = 4

        while current_date <= end_date:
            # Получаем уроки на этот день
            day_lessons = [lesson for lesson in lessons if lesson.date == current_date]

            if not day_lessons:
                # Если нет уроков, создаем пустую строку
                date_text = f"{self._get_weekday_english(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                date_cell = worksheet.cell(row=row, column=1, value=date_text)
                date_cell.font = header_font
                date_cell.border = border
                date_cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

                # Заполняем пустые временные слоты
                for col, time_slot in enumerate(time_slots, 2):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = border

                # Пустая колонка групп
                groups_cell = worksheet.cell(row=row, column=group_col, value="")
                groups_cell.border = border

                row += 1
            else:
                # Определяем уникальные группы для этого дня
                groups_for_day = {}
                for lesson in day_lessons:
                    if lesson.group:
                        group_name = lesson.group.name
                        if group_name not in groups_for_day:
                            groups_for_day[group_name] = []
                        groups_for_day[group_name].append(lesson)

                # Если есть только одна группа или уроки без группы
                if len(groups_for_day) <= 1:
                    # Стандартная строка
                    date_text = f"{self._get_weekday_english(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                    date_cell = worksheet.cell(row=row, column=1, value=date_text)
                    date_cell.font = header_font
                    date_cell.border = border
                    date_cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                    # Заполняем временные слоты
                    for col, time_slot in enumerate(time_slots, 2):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = border

                        # Ищем урок в этом временном слоте
                        lesson_in_slot = self._find_lesson_in_slot(
                            day_lessons, time_slot
                        )
                        if lesson_in_slot:
                            self._fill_lesson_cell(cell, lesson_in_slot)

                    # Заполняем колонку групп
                    groups_text = (
                        "\n".join(sorted(groups_for_day.keys()))
                        if groups_for_day
                        else ""
                    )
                    groups_cell = worksheet.cell(
                        row=row, column=group_col, value=groups_text
                    )
                    groups_cell.font = cell_font
                    groups_cell.border = border
                    groups_cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                    row += 1
                else:
                    # Несколько групп - создаем отдельные строки для каждой группы
                    start_row = row
                    num_groups = len(groups_for_day)

                    # Создаем объединенную ячейку для даты
                    date_text = f"{self._get_weekday_english(current_date)}\n{current_date.strftime('%d.%m.%Y')}"
                    worksheet.merge_cells(f"A{start_row}:A{start_row + num_groups - 1}")
                    date_cell = worksheet.cell(row=start_row, column=1, value=date_text)
                    date_cell.font = header_font
                    date_cell.border = border
                    date_cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )

                    # Создаем строки для каждой группы
                    for group_name in sorted(groups_for_day.keys()):
                        group_lessons = groups_for_day[group_name]

                        # Заполняем временные слоты только для этой группы
                        for col, time_slot in enumerate(time_slots, 2):
                            cell = worksheet.cell(row=row, column=col)
                            cell.border = border

                            # Ищем урок в этом временном слоте для этой группы
                            lesson_in_slot = self._find_lesson_in_slot(
                                group_lessons, time_slot
                            )
                            if lesson_in_slot:
                                self._fill_lesson_cell(cell, lesson_in_slot)

                        # Заполняем колонку групп
                        groups_cell = worksheet.cell(
                            row=row, column=group_col, value=group_name
                        )
                        groups_cell.font = cell_font
                        groups_cell.border = border
                        groups_cell.alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True
                        )

                        row += 1

            current_date += timedelta(days=1)

        # Устанавливаем ширину колонок
        worksheet.column_dimensions["A"].width = 15  # Дата
        for col in range(2, len(time_slots) + 2):
            worksheet.column_dimensions[
                get_column_letter(col)
            ].width = 20  # Временные слоты
        worksheet.column_dimensions[get_column_letter(group_col)].width = 15  # Группа

        # Устанавливаем высоту строк
        for row_num in range(4, row):
            worksheet.row_dimensions[row_num].height = 60

    def _get_weekday_english(self, date_obj) -> str:
        """
        Get the English uppercase weekday name for a date.

        Args:
            date_obj (date): Target date.

        Returns:
            str: Weekday name (e.g., "MONDAY").
        """
        """Получение дня недели на английском"""
        weekdays = {
            0: "MONDAY",
            1: "TUESDAY",
            2: "WEDNESDAY",
            3: "THURSDAY",
            4: "FRIDAY",
            5: "SATURDAY",
            6: "SUNDAY",
        }
        return weekdays[date_obj.weekday()]

    def _find_lesson_in_slot(
        self, lessons: List[Lesson], time_slot: str
    ) -> Optional[Lesson]:
        """
        Find a lesson that overlaps a given time slot.

        Overlap conditions (inclusive):
        - Lesson starts before slot end and ends after slot start, or
        - Either endpoint falls within the other interval, or
        - One interval fully contains the other.

        Args:
            lessons (List[Lesson]): Candidate lessons for a specific date.
            time_slot (str): Slot label "HH:MM-HH:MM".

        Returns:
            Optional[Lesson]: Matching lesson if found, otherwise None.
        """
        """Поиск урока в временном слоте"""
        slot_start, slot_end = time_slot.split("-")
        slot_start_time = datetime.strptime(slot_start, "%H:%M").time()
        slot_end_time = datetime.strptime(slot_end, "%H:%M").time()

        for lesson in lessons:
            # Проверяем пересечение времени
            if (
                lesson.start_time <= slot_start_time < lesson.end_time
                or lesson.start_time < slot_end_time <= lesson.end_time
                or (
                    lesson.start_time >= slot_start_time
                    and lesson.end_time <= slot_end_time
                )
            ):
                return lesson

        return None

    def _fill_lesson_cell(self, cell, lesson: Lesson):
        """
        Fill an Excel cell with lesson data and apply styling.

        Content:
        - Subject name (first line)
        - Professor full name (second line)
        - Delivery: "ONLINE" or "Room: <number>"

        Styling:
        - Center alignment, white bold font
        - Background color from subject color (hex) or gray fallback

        Args:
            cell (openpyxl.cell.Cell): Target cell to fill.
            lesson (Lesson): Lesson providing values and color.

        Returns:
            None
        """
        """Заполнение ячейки урока"""
        # Формируем текст урока
        lesson_text = f"{lesson.subject.name}\n"
        lesson_text += f"{lesson.professor.name} {lesson.professor.surname}\n"

        if lesson.is_online:
            lesson_text += "ONLINE"
        elif lesson.room:
            lesson_text += f"Room: {lesson.room.number}"

        cell.value = lesson_text
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.font = Font(size=9, color="FFFFFF", bold=True)

        # Устанавливаем цвет фона из цвета предмета
        if lesson.subject.color:
            # Убираем # из hex цвета
            color_hex = lesson.subject.color.replace("#", "")
            cell.fill = PatternFill(
                start_color=color_hex, end_color=color_hex, fill_type="solid"
            )
        else:
            # Цвет по умолчанию (серый)
            cell.fill = PatternFill(
                start_color="6b7280", end_color="6b7280", fill_type="solid"
            )
